import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import copy
from osdagbridge.desktop.ui.dialogs.generate_results_values_builder import (
    resolve_bridge_config_summary,resolve_material_properties_steel
)

from PySide6.QtWidgets import ( 
    QDialog, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QTreeWidget, QTreeWidgetItem, QComboBox, QSizeGrip, QFileDialog,
    QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QSizePolicy, QStackedWidget, QStyledItemDelegate, QApplication
)
from PySide6.QtCore import Qt, QRect
from PySide6.QtGui import QColor, QFont, QPainter, QPainterPath, QPen
from osdagbridge.core.bridge_types.plate_girder.ui_fields_additional_input import GENERATE_RESULTS_DEFAULTS
from osdagbridge.desktop.ui.utils.custom_titlebar import CustomTitleBar
from osdagbridge.desktop.ui.dialogs.custom_messagebox import CustomMessageBox, MessageBoxType
from osdagbridge.desktop.ui.dialogs.generate_results_values_builder import resolve_table
from osdagbridge.desktop.ui.utils.custom_cursors import pointing_hand_cursor


class CheckboxDelegate(QStyledItemDelegate):
    """Custom delegate to paint checkmarks/partial marks inside checkbox boxes."""

    def paint(self, painter, option, index):
        super().paint(painter, option, index)

        tree = self.parent()
        if not isinstance(tree, QTreeWidget):
            return

        item = tree.itemFromIndex(index)
        if not item:
            return

        check_state = item.checkState(0)

        if check_state == Qt.Checked:
            symbol = "✓"
        elif check_state == Qt.PartiallyChecked:
            symbol = "−"
        else:
            return

        checkbox_rect = QRect(
            option.rect.left() + 4,
            option.rect.top(),
            20,
            option.rect.height()
        )

        painter.setPen(QColor("#90AF13"))
        painter.setFont(QFont("Arial", 10))
        painter.drawText(checkbox_rect, Qt.AlignCenter, symbol)

class NoScrollComboBox(QComboBox):
    def wheelEvent(self, event):
        event.ignore()


def apply_field_style(widget):
    widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
    widget.setMinimumHeight(30)

    if isinstance(widget, QComboBox):
        widget.setStyleSheet("""
            QComboBox{
                padding: 1px 7px;
                border: 1px solid black;
                border-radius: 5px;
                background-color: white;
                color: black;
            }
            QComboBox::drop-down{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                border-left: 0px;
            }
            QComboBox::down-arrow{
                image: url(:/osdagbridge/vectors/arrow_down_light.svg);
                width: 20px;
                height: 20px;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView{
                background-color: white;
                border: 1px solid black;
                color: black;
            }
            QComboBox:disabled{
                padding: 1px 7px;
                border: 1px solid #666;
                border-radius: 5px;
                background-color: #f1f1f1;
                color: #666;
            }
        """)

class RoundedTableFrame(QFrame):

    RADIUS       = 12
    BORDER_COLOR = QColor("#90AF13")
    HEADER_COLOR = QColor("#90AF13")
    HEADER_H     = 34

    def __init__(self, parent=None):
        super().__init__(parent)
        self._greyed = True
        self.setObjectName("roundedTableFrame")
        self.setStyleSheet(
            "QFrame#roundedTableFrame { border: none; background: transparent; }"
        )

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        r  = self.RADIUS
        w  = self.width()
        h  = self.height()
        hh = self.HEADER_H

        full_rect = self.rect().toRectF().adjusted(0.5, 0.5, -0.5, -0.5)
        full_path = QPainterPath()
        full_path.addRoundedRect(full_rect, r, r)

        painter.fillPath(full_path, QColor("white"))

        header_rect = full_rect.adjusted(0, 0, 0, -(h - hh - 1))
        header_path = QPainterPath()
        header_path.addRoundedRect(header_rect, r, r)
        square_bottom = QPainterPath()
        square_bottom.addRect(full_rect.adjusted(0, r, 0, -(h - hh - 1)))
        header_path = header_path.united(square_bottom)

        active_color = QColor("#cccccc") if self._greyed else self.HEADER_COLOR
        border_color = QColor("#cccccc") if self._greyed else self.BORDER_COLOR
        painter.fillPath(header_path, active_color)
        pen = QPen(border_color, 1.0)

        pen = QPen(self.BORDER_COLOR, 1.0)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)
        painter.setBrush(Qt.NoBrush)
        painter.drawPath(full_path)

        painter.end()

    def set_greyed(self, greyed: bool):
        self._greyed = greyed
        self.update()    

class VerticalLabel(QWidget):
    """Paints text rotated 90° counter-clockwise, centered in the widget."""

    def __init__(self, text="", color="#90AF13", font_size=13, font_weight=700, parent=None):
        super().__init__(parent)
        self._text        = text
        self._color       = QColor(color)
        self._font_size   = font_size
        self._font_weight = font_weight
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def setText(self, text):
        self._text = text
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.TextAntialiasing)

        font = QFont("Arial", self._font_size)
        font.setWeight(QFont.Weight(self._font_weight))
        font.setLetterSpacing(QFont.AbsoluteSpacing, 3)
        painter.setFont(font)
        painter.setPen(self._color)

        painter.translate(self.width() / 2, self.height() / 2)
        painter.rotate(-90)
        painter.drawText(
            QRect(-self.height() // 2, -self.width() // 2, self.height(), self.width()),
            Qt.AlignCenter,
            self._text
        )
        painter.end()

class GenerateResultsPage(QWidget):
    """Page 1 — table selection + load case/member combos."""

    def __init__(self, on_show_selections, parent=None):
        super().__init__(parent)
        self._on_show_selections = on_show_selections
        self._setup_ui()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(18, 18, 18, 14)
        main_layout.setSpacing(12)

        body = QHBoxLayout()
        body.setSpacing(14)

        left_card = QFrame()
        left_card.setObjectName("leftCard")
        left_card.setStyleSheet("""
            QFrame#leftCard {
                background: white;
                border: 1px solid #90af13;
                border-radius: 10px;
            }
        """)

        left_layout = QVBoxLayout(left_card)
        left_layout.setContentsMargins(14, 12, 14, 12)
        left_layout.setSpacing(12)

        top_bar = QHBoxLayout()

        title = QLabel("Select Tables")
        title.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: 700;
                color: #2d2d2d;
            }
        """)
        top_bar.addWidget(title)
        top_bar.addStretch()

        action_btn_style = """
            QPushButton {
                background: white;
                color: #3a3a3a;
                border: 1px solid #90af13;
                border-radius: 6px;
                padding: 4px 12px;
                font-size: 12px;
                font-weight: 500;
            }
            QPushButton:hover {
                background: #90AF13;
                border-color: #90AF13;
                color: white;
            }
            QPushButton:pressed {
                background: #7a9a10;
                border-color: #7a9a10;
                color: white;
            }
        """

        self.select_btn = QPushButton("Select All")
        self.select_btn.setStyleSheet(action_btn_style)
        self.select_btn.clicked.connect(self.select_all_items)

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setStyleSheet(action_btn_style)
        self.clear_btn.clicked.connect(self.clear_all_items)

        top_bar.addWidget(self.select_btn)
        top_bar.addWidget(self.clear_btn)

        left_layout.addLayout(top_bar)

        # Tree
        self.tree = QTreeWidget()
        self.tree.setStyleSheet("""
            QTreeWidget {
                background: white;
                border: none;
                color: black;
                font-size: 12px;
                outline: 0;
                show-decoration-selected: 0;
            }
            QTreeWidget::item {
                height: 24px;
                border: none;
                padding: 2px 4px;
            }
            QTreeWidget::item:hover {
                background: rgba(144,175,19,45);
                border: none;
                color: black;
            }
            QTreeWidget::indicator {
                width: 14px;
                height: 14px;
            }
            QTreeWidget::indicator:unchecked {
                background-color: white;
                border: 1px solid #90af13;
                border-radius: 3px;
            }
            QTreeWidget::indicator:checked {
                background-color: white;
                border: 1px solid #90af13;
                border-radius: 3px;
            }
            QTreeWidget::indicator:indeterminate {
                background-color: white;
                border: 1px solid #90af13;
                border-radius: 3px;
            }
            QTreeWidget::item:selected {
                background: rgba(144,175,19,55);
                border: none;
                color: black;
            }
            QTreeWidget::branch:selected {
                background: transparent;
            }
            QTreeView::branch {
                background: transparent;
                border: none;
                image: none;
            }
            QTreeView::branch:selected {
                background: transparent;
            }
            QTreeView::branch:has-siblings:!adjoins-item,
            QTreeView::branch:has-siblings:adjoins-item,
            QTreeView::branch:has-children:!has-siblings:closed,
            QTreeView::branch:has-children:!has-siblings:open,
            QTreeView::branch:closed:has-children,
            QTreeView::branch:open:has-children {
                border-image: none;
                image: none;
            }
            QScrollArea {
                background: transparent;
            }

            QScrollBar:horizontal {
                background: #E0E0E0;
                height: 8px;
                border-radius: 12px;
            }

            QScrollBar::handle:horizontal {
                background: #A0A0A0;
                min-width: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:horizontal:hover {
                background: #707070;
            }

            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                width: 0px;
            }

            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: transparent;
            }

            QScrollBar:vertical {
                background: #E0E0E0;
                width: 8px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical {
                background: #A0A0A0;
                min-height: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical:hover {
                background: #707070;
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: transparent;
            }

        """)
        self.tree.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tree.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        self.tree.setVerticalScrollMode(QTreeWidget.ScrollPerPixel)
        self.tree.setHorizontalScrollMode(QTreeWidget.ScrollPerPixel)
        self.tree.setItemDelegate(CheckboxDelegate(self.tree))
        self.tree.itemClicked.connect(self._on_tree_item_clicked)
        self.tree.setFocusPolicy(Qt.NoFocus)
        self.tree.setAllColumnsShowFocus(False)
        self.tree.setRootIsDecorated(False)
        self.tree.setIndentation(18)
        self.tree.setHeaderHidden(True)
        self.tree.itemChanged.connect(self._handle_item_changed)

        left_layout.addWidget(self.tree, 1)

        body.addWidget(left_card, 2)

        # ── RIGHT CARD ─────────────────────────────────────────────────────
        right_card = QFrame()
        right_card.setObjectName("rightCard")
        right_card.setStyleSheet("""
            QFrame#rightCard {
                background: white;
                border: 1px solid #90AF13;
                border-radius: 10px;
            }
        """)

        right_layout = QVBoxLayout(right_card)
        right_layout.setContentsMargins(14, 12, 14, 12)
        right_layout.setSpacing(10)

        lc_label = QLabel("Load Case")
        lc_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: 700;
                color: #2d2d2d;
            }
        """)

        member_label = QLabel("Member")
        member_label.setStyleSheet(lc_label.styleSheet())
        right_layout.addWidget(member_label)

        self.member_combo = NoScrollComboBox()
        self.member_combo.addItems([
            "All Girders",
            "Girder 1",
            "Girder 2",
            "Girder 3",
            "Girder 4"
        ])
        apply_field_style(self.member_combo)
        right_layout.addWidget(self.member_combo)

        right_layout.addWidget(lc_label)

        self.load_case_combo = NoScrollComboBox()
        self.load_case_combo.addItems([
            "Dead Load", "Wering Surface Load", "Secondary impact Dead Load", "Live Load",
            "Wind Load", "Seismic Load", "Temperature Load",
            "ULS Combo", "SLS Combo"
        ])
        apply_field_style(self.load_case_combo)
        right_layout.addWidget(self.load_case_combo)

        right_layout.addStretch()

        body.addWidget(right_card, 1)

        main_layout.addLayout(body)

        # ── FOOTER ─────────────────────────────────────────────────────────
        footer = QHBoxLayout()
        footer.addStretch()

        btn_style = """
            QPushButton {
                background: white;
                color: black;
                border: 1px solid #c0c0c0;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
                min-width: 110px;
            }
            QPushButton:hover {
                background: #90AF13;
                border-color: #90AF13;
                color: white;
            }
        """

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(btn_style)
        self._cancel_btn = cancel_btn

        self.show_btn = QPushButton("Show Selections")
        self.show_btn.setStyleSheet(btn_style)
        self.show_btn.clicked.connect(self._on_show_selections)

        footer.addWidget(cancel_btn)
        footer.addWidget(self.show_btn)

        main_layout.addLayout(footer)

        self._build_tree()

    # ── Tree Build ──────────────────────────────────────────────────────────
    def _build_tree(self):
        self.tree.blockSignals(True)

        for main_key, main_val in GENERATE_RESULTS_DEFAULTS.items():
            parent = QTreeWidgetItem(self.tree)
            parent.setText(0, main_val["label"])
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.Unchecked)

            for sub_key, sub_val in main_val.items():
                if sub_key in ("id", "label"):
                    continue

                child = QTreeWidgetItem(parent)
                child.setText(0, sub_val["label"])
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                child.setCheckState(0, Qt.Unchecked)

                for table_key, table_data in sub_val.items():
                    if table_key in ("id", "label"):
                        continue

                    leaf = QTreeWidgetItem(child)
                    leaf.setText(0, table_data["label"])
                    leaf.setFlags(leaf.flags() | Qt.ItemIsUserCheckable)
                    leaf.setCheckState(0, Qt.Unchecked)

        self.tree.expandToDepth(1)
        self.tree.blockSignals(False)

    # ── Checkbox Logic ──────────────────────────────────────────────────────
    def _handle_item_changed(self, item, column):
        state = item.checkState(0)
        self.tree.blockSignals(True)
        for i in range(item.childCount()):
            child = item.child(i)
            child.setCheckState(0, state)
            self._update_children(child, state)
        self._update_parents(item)
        self.tree.blockSignals(False)

    def _update_children(self, item, state):
        for i in range(item.childCount()):
            child = item.child(i)
            child.setCheckState(0, state)
            self._update_children(child, state)

    def _update_parents(self, item):
        parent = item.parent()
        while parent:
            checked   = 0
            unchecked = 0
            for i in range(parent.childCount()):
                st = parent.child(i).checkState(0)
                if st == Qt.Checked:
                    checked += 1
                elif st == Qt.Unchecked:
                    unchecked += 1
            if checked == parent.childCount():
                parent.setCheckState(0, Qt.Checked)
            elif unchecked == parent.childCount():
                parent.setCheckState(0, Qt.Unchecked)
            else:
                parent.setCheckState(0, Qt.PartiallyChecked)
            parent = parent.parent()

    # ── Select / Clear All ──────────────────────────────────────────────────
    def select_all_items(self):
        self.tree.blockSignals(True)
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            item.setCheckState(0, Qt.Checked)
            self._update_children(item, Qt.Checked)
        self.tree.blockSignals(False)

    def clear_all_items(self):
        self.tree.blockSignals(True)
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            item.setCheckState(0, Qt.Unchecked)
            self._update_children(item, Qt.Unchecked)
        self.tree.blockSignals(False)

    # ── Selection ───────────────────────────────────────────────────────────
    def get_selected_tables(self):
        selected = []
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            main_item = root.child(i)
            for j in range(main_item.childCount()):
                group = main_item.child(j)
                for k in range(group.childCount()):
                    leaf = group.child(k)
                    if leaf.checkState(0) == Qt.Checked:
                        selected.append(leaf.text(0))
        return selected

    # ── Tree Click Toggle ───────────────────────────────────────────────────
    def _on_tree_item_clicked(self, item, column):
        pos = self.tree.viewport().mapFromGlobal(
            self.tree.viewport().cursor().pos()
        )
        index        = self.tree.indexAt(pos)
        checkbox_rect = self.tree.visualRect(index)
        checkbox_rect.setWidth(20)

        if checkbox_rect.contains(pos):
            return

        current = item.checkState(0)
        if current == Qt.Checked:
            item.setCheckState(0, Qt.Unchecked)
        else:
            item.setCheckState(0, Qt.Checked)


class ExportTablePage(QWidget):
    """Page 2 — tree of selected tables + preview + Excel export."""

    _COL_MIN_WIDTH = 120
    _COL_MAX_WIDTH = 320

    def __init__(self, on_back, on_cancel, parent=None):
        super().__init__(parent)
        self._on_back    = on_back
        self._on_cancel  = on_cancel
        self._is_expanded = True
        self._setup_ui()

    # ── resizeEvent ─────────────────────────────────────────────────────────
    def resizeEvent(self, event):
        """
        Called whenever the widget (and therefore its viewport) changes size —
        triggered by dialog resize OR panel collapse/expand.
        We re-run the proportional fill so columns always use the available space.
        """
        super().resizeEvent(event)
        self._fit_columns()

    # ── _fit_columns ────────────────────────────────────────────────────────
    def _fit_columns(self):
        """
        Proportionally scale column widths to fill the full viewport width
        while preserving the relative width ratios set by content or by the
        user when they manually drag a column divider.

        Design decisions
        ────────────────
        • We never stamp every column to the same fixed pixel value.
        • We never divide viewport_w equally across columns.
        • We only *scale up* when there is leftover space; we never shrink
          below the natural content width (that would hide data).
        • If total content width already exceeds the viewport the horizontal
          scrollbar appears naturally — we do not force-shrink anything.
        • The per-column minimum (_COL_MIN_WIDTH) prevents a column from
          becoming invisible after extreme scaling.
        """
        table     = self.preview
        col_count = table.columnCount()
        if col_count == 0:
            return

        viewport_w = table.viewport().width()
        if viewport_w <= 0:
            return

        current_total = sum(table.columnWidth(i) for i in range(col_count))
        if current_total <= 0:
            return

        if current_total < viewport_w:
            scale = viewport_w / current_total
            for i in range(col_count):
                new_w = int(table.columnWidth(i) * scale)
                table.setColumnWidth(i, max(self._COL_MIN_WIDTH, new_w))

    # ── _setup_ui ───────────────────────────────────────────────────────────
    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(18, 18, 18, 14)
        main_layout.setSpacing(14)

        card        = QFrame()
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(10)

        # ── MAIN SPLIT ─────────────────────────────────────────────────────
        main_split = QHBoxLayout()
        main_split.setSpacing(0)
        main_split.setContentsMargins(0, 0, 0, 0)

        # ── LEFT PANEL ─────────────────────────────────────────────────────
        self.left_box = QFrame()
        self.left_box.setObjectName("leftBox")
        self.left_box.setMinimumWidth(320)
        self.left_box.setMaximumWidth(400)
        self.left_box.setStyleSheet("""
            QFrame#leftBox {
                border: 1px solid #90AF13;
                border-radius: 12px;
                background: white;
            }
        """)

        left_layout = QVBoxLayout(self.left_box)
        left_layout.setContentsMargins(8, 10, 8, 8)
        left_layout.setSpacing(8)

        action_btn_style = """
            QPushButton {
                background: white;
                color: #3a3a3a;
                border: 1px solid #90af13;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: 500;
            }
            QPushButton:hover {
                background: #90AF13;
                border-color: #90AF13;
                color: white;
            }
            QPushButton:pressed {
                background: #7a9a10;
                border-color: #7a9a10;
                color: white;
            }
        """

        back_btn_style = """
            QPushButton {
                background: white;
                color: #3a3a3a;
                border: 1px solid #90af13;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: #90AF13;
                border-color: #90AF13;
                color: white;
            }
            QPushButton:pressed {
                background: #7a9a10;
                border-color: #7a9a10;
                color: white;
            }
        """

        top_bar = QHBoxLayout()
        top_bar.setSpacing(6)

        self.back_btn = QPushButton("← Back")
        self.back_btn.setStyleSheet(back_btn_style)
        self.back_btn.clicked.connect(self._on_back)
        top_bar.addWidget(self.back_btn)

        top_bar.addStretch()

        self.select_btn = QPushButton("Select All")
        self.select_btn.setStyleSheet(action_btn_style)
        self.select_btn.clicked.connect(self.select_all_items)

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setStyleSheet(action_btn_style)
        self.clear_btn.clicked.connect(self.clear_all_items)

        top_bar.addWidget(self.select_btn)
        top_bar.addWidget(self.clear_btn)

        left_layout.addLayout(top_bar)

        # Tree
        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.setStyleSheet("""
            QTreeWidget {
                border: none;
                background: white;
                outline: none;
                font-size: 12px;
                color: black;
            }
            QTreeWidget::item {
                height: 28px;
                padding: 2px 4px;
                border: none;
            }
            QTreeWidget::item:hover {
                background: rgba(144, 175, 19, 45);
            }
            QTreeWidget::item:selected {
                background: rgba(144, 175, 19, 60);
                color: black;
            }
            QTreeWidget::indicator {
                width: 14px;
                height: 14px;
            }
            QTreeWidget::indicator:unchecked {
                background-color: white;
                border: 1px solid #90AF13;
                border-radius: 3px;
            }
            QTreeWidget::indicator:checked {
                background-color: white;
                border: 1px solid #90AF13;
                border-radius: 3px;
            }
            QTreeWidget::indicator:indeterminate {
                background-color: white;
                border: 1px solid #90AF13;
                border-radius: 3px;
            }
            QScrollArea {
                background: transparent;
            }
            QScrollBar:horizontal {
                background: #E0E0E0;
                height: 8px;
                border-radius: 12px;
            }

            QScrollBar::handle:horizontal {
                background: #A0A0A0;
                min-width: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:horizontal:hover {
                background: #707070;
            }

            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                width: 0px;
            }

            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: transparent;
            }
            QScrollBar:vertical {
                background: #E0E0E0;
                width: 8px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical {
                background: #A0A0A0;
                min-height: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical:hover {
                background: #707070;
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: transparent;
            }
        """)
        self.tree.setItemDelegate(CheckboxDelegate(self.tree))
        self.tree.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tree.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tree.setHorizontalScrollMode(QTreeWidget.ScrollPerPixel)
        self.tree.setWordWrap(False)
        self.tree.setUniformRowHeights(False)
        self.tree.setIndentation(20)
        self.tree.itemClicked.connect(self._load_selected_table)
        self.tree.itemChanged.connect(self._handle_item_changed)

        left_layout.addWidget(self.tree, 1)

        self.vertical_label = VerticalLabel(
            text="SELECT TABLES",
            color="#000000",
            font_size=8,
            font_weight=200
        )
        self.vertical_label.hide()
        left_layout.addWidget(self.vertical_label, 1)

        self.left_container = QWidget()
        self.left_container.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        left_container_layout = QHBoxLayout(self.left_container)
        left_container_layout.setContentsMargins(0, 0, 8, 0)
        left_container_layout.setSpacing(0)
        left_container_layout.addWidget(self.left_box, 1)

        self._toggle_btn_style_expanded = """
            QPushButton {
                background-color: #90AF13;
                color: white;
                border: none;
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
                border-top-left-radius: 0px;
                border-bottom-left-radius: 0px;
                font-size: 18px;
                font-weight: bold;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover { background-color: #7a9a10; }
            QPushButton:pressed { background-color: #6a8a0e; }
        """
        self._toggle_btn_style_collapsed = """
            QPushButton {
                background-color: #90AF13;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover { background-color: #7a9a10; }
            QPushButton:pressed { background-color: #6a8a0e; }
        """

        self.toggle_btn = QPushButton("‹")
        self.toggle_btn.setFixedSize(20, 56)
        self.toggle_btn.setCursor(pointing_hand_cursor())
        self.toggle_btn.clicked.connect(self._toggle_tree_panel)
        self.toggle_btn.setStyleSheet(self._toggle_btn_style_expanded)
        left_container_layout.addWidget(self.toggle_btn, 0, Qt.AlignVCenter)

        main_split.addWidget(self.left_container)

        # ── RIGHT PANEL ─────────────────────────────────────
        self.right_box = RoundedTableFrame()

        right_layout = QVBoxLayout(self.right_box)
        right_layout.setContentsMargins(1, 0, 1, 1)
        right_layout.setSpacing(0)

        self.preview = QTableWidget()
        self.preview.setColumnCount(2)
        self.preview.setHorizontalHeaderLabels(["Parameter", "Value"])

        header = self.preview.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(self._COL_MIN_WIDTH)
        header.setSectionsClickable(False)
        header.setStretchLastSection(False)         
        header.setFixedHeight(34)

        self.preview.verticalHeader().setVisible(False)
        self.preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)   
        self.preview.setFrameShape(QFrame.NoFrame)
        self.preview.setShowGrid(True)
        self.preview.setStyleSheet("""
            QTableWidget {
                border: none;
                background: transparent;
                gridline-color: #d0d0d0;
            }
            QTableWidget::item {
                padding: 4px;
                color: black;
                background: white;
            }
            QTableWidget::item:selected {
                color: black;
                background: rgba(144, 175, 19, 60);
            }
            QHeaderView {
                border: none;
                background: transparent;
            }
            QHeaderView::section {
                background: transparent;
                color: white;
                border: none;
                border-right: 1px solid rgba(255,255,255,0.25);
                padding: 6px;
                font-weight: 600;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QTableCornerButton::section {
                border: none;
                background: transparent;
            }
            QScrollArea {
                background: transparent;
            }
            QScrollBar:horizontal {
                background: #E0E0E0;
                height: 8px;
                border-radius: 12px;
                margin: 0px 20px 0px 20px;
            }

            QScrollBar::handle:horizontal {
                background: #A0A0A0;
                min-width: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:horizontal:hover {
                background: #707070;
            }

            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                width: 0px;
            }

            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: transparent;
            }

            QScrollBar:vertical {
                background: #E0E0E0;
                width: 8px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical {
                background: #A0A0A0;
                min-height: 30px;
                border-radius: 12px;
            }

            QScrollBar::handle:vertical:hover {
                background: #707070;
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: transparent;
            }
            """)

        # ── Placeholder───────────────────
        self.placeholder = QWidget()
        self.placeholder.setObjectName("previewPlaceholder")
        self.placeholder.setStyleSheet(
            "QWidget#previewPlaceholder { background: transparent; }"
        )
        ph_layout = QVBoxLayout(self.placeholder)
        ph_layout.setAlignment(Qt.AlignCenter)
        ph_layout.setSpacing(8)

        ph_text = QLabel("Select a table to preview values")
        ph_text.setAlignment(Qt.AlignCenter)
        ph_text.setStyleSheet("""
            QLabel {
                color: #aaaaaa;
                font-size: 13px;
                font-weight: 500;
                background: transparent;
                border: none;
            }
        """)

        ph_layout.addWidget(ph_text)

        self.preview_stack = QStackedWidget()
        self.preview_stack.addWidget(self.placeholder)
        self.preview_stack.addWidget(self.preview)
        self.preview_stack.setCurrentIndex(0)

        right_layout.addWidget(self.preview_stack, 1)

        main_split.addWidget(self.right_box, 1)

        card_layout.addLayout(main_split)
        main_layout.addWidget(card, 1)

        # ── FOOTER ─────────────────────────────────────────────────────────
        footer = QHBoxLayout()
        footer.addStretch()

        btn_style = """
            QPushButton {
                background: white;
                color: black;
                border: 1px solid #c0c0c0;
                border-radius: 6px;
                padding: 8px 18px;
                font-weight: 600;
                min-width: 110px;
            }
            QPushButton:hover {
                background: #90AF13;
                border-color: #90AF13;
                color: white;
            }
        """

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(btn_style)
        cancel_btn.clicked.connect(self._on_cancel)

        export_btn = QPushButton("Export")
        export_btn.setStyleSheet(btn_style)
        export_btn.clicked.connect(self._export_excel)

        footer.addWidget(cancel_btn)
        footer.addWidget(export_btn)

        main_layout.addLayout(footer)

    # ── Load Data into Page ─────────────────────────────────────────────────
    def load_data(self, selected_tables: dict):
        """Populate tree with the export_data dict passed from page 1."""
        self.tree.blockSignals(True)
        self.tree.clear()

        for lvl1, sub1 in selected_tables.items():
            p1 = QTreeWidgetItem(self.tree, [lvl1])
            p1.setFlags(p1.flags() | Qt.ItemIsUserCheckable)
            p1.setCheckState(0, Qt.Unchecked)

            for lvl2, sub2 in sub1.items():
                p2 = QTreeWidgetItem(p1, [lvl2])
                p2.setFlags(p2.flags() | Qt.ItemIsUserCheckable)
                p2.setCheckState(0, Qt.Unchecked)

                for lvl3, data in sub2.items():
                    p3 = QTreeWidgetItem(p2, [lvl3])
                    p3.setFlags(p3.flags() | Qt.ItemIsUserCheckable)
                    p3.setCheckState(0, Qt.Unchecked)
                    p3.setData(0, Qt.UserRole, data)

        self.tree.expandAll()
        self.tree.resizeColumnToContents(0)
        self.tree.updateGeometry()
        self.tree.blockSignals(False)

        # Reset preview
        self.preview.clear()
        self.preview.setColumnCount(2)
        self.preview.setHorizontalHeaderLabels(["Parameter", "Value"])
        self.preview.setRowCount(0)
        self._set_preview_empty()

    # ── Left Panel Toggle ───────────────────────────────────────────────────
    def _toggle_tree_panel(self):
        if self._is_expanded:
            self._is_expanded = False
            self.left_box.setMinimumWidth(0)
            self.left_box.setMaximumWidth(60)
            self.left_box.setFixedWidth(30)
            self.toggle_btn.setText("›")
            self.toggle_btn.setStyleSheet(self._toggle_btn_style_collapsed)
            self.select_btn.hide()
            self.clear_btn.hide()
            self.back_btn.hide()
            self.tree.hide()
            self.vertical_label.show()
        else:
            self._is_expanded = True
            self.left_box.setMinimumWidth(320)
            self.left_box.setMaximumWidth(400)
            self.toggle_btn.setText("‹")
            self.toggle_btn.setStyleSheet(self._toggle_btn_style_expanded)
            self.vertical_label.hide()
            self.tree.show()
            self.select_btn.show()
            self.clear_btn.show()
            self.back_btn.show()

        QApplication.processEvents()
        self._fit_columns()

    # ── Select / Clear All ──────────────────────────────────────────────────
    def select_all_items(self):
        self.tree.blockSignals(True)
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            item.setCheckState(0, Qt.Checked)
            self._update_children(item, Qt.Checked)
        self.tree.blockSignals(False)

    def clear_all_items(self):
        self.tree.blockSignals(True)
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            item.setCheckState(0, Qt.Unchecked)
            self._update_children(item, Qt.Unchecked)
        self.tree.blockSignals(False)

    # ── Table Preview ───────────────────────────────────────────────────────
    def _load_selected_table(self, item, column):
        data = item.data(0, Qt.UserRole)
        if not isinstance(data, dict):
            self._set_preview_empty()
            return

        columns = data.get("columns", [])
        rows    = data.get("rows",    [])
        if not columns or not rows:
            self._set_preview_empty()
            return

        self.right_box.set_greyed(False)
        self.preview_stack.setCurrentIndex(1)

        self.preview.clear()
        self.preview.setColumnCount(len(columns))
        self.preview.setRowCount(len(rows))
        self.preview.setHorizontalHeaderLabels([str(c) for c in columns])

        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                cell = QTableWidgetItem(str(val))
                cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                cell.setForeground(Qt.black)
                cell.setBackground(QColor("#ffffff"))
                cell.setTextAlignment(Qt.AlignCenter)
                self.preview.setItem(r_idx, c_idx, cell)

        self.preview.verticalHeader().setVisible(False)
        self.preview.resizeColumnsToContents()

        for i in range(self.preview.columnCount()):
            w = self.preview.columnWidth(i)
            self.preview.setColumnWidth(
                i, max(self._COL_MIN_WIDTH, min(w, self._COL_MAX_WIDTH))
            )

        self._fit_columns()

    def _set_preview_empty(self):
        """Show the placeholder and grey out the right panel."""
        self.right_box.set_greyed(True)
        self.preview_stack.setCurrentIndex(0)

    # ── Checkbox Logic ──────────────────────────────────────────────────────
    def _handle_item_changed(self, item, column):
        state = item.checkState(0)
        self.tree.blockSignals(True)
        for i in range(item.childCount()):
            child = item.child(i)
            child.setCheckState(0, state)
            self._update_children(child, state)
        self._update_parents(item)
        self.tree.blockSignals(False)

    def _update_children(self, item, state):
        for i in range(item.childCount()):
            child = item.child(i)
            child.setCheckState(0, state)
            self._update_children(child, state)

    def _update_parents(self, item):
        parent = item.parent()
        while parent:
            checked   = 0
            unchecked = 0
            for i in range(parent.childCount()):
                st = parent.child(i).checkState(0)
                if st == Qt.Checked:
                    checked += 1
                elif st == Qt.Unchecked:
                    unchecked += 1
            if checked == parent.childCount():
                parent.setCheckState(0, Qt.Checked)
            elif unchecked == parent.childCount():
                parent.setCheckState(0, Qt.Unchecked)
            else:
                parent.setCheckState(0, Qt.PartiallyChecked)
            parent = parent.parent()

    # ── Get Checked Tables ──────────────────────────────────────────────────
    def get_checked_tables(self):
        checked_items = []
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            lvl1 = root.child(i)
            for j in range(lvl1.childCount()):
                lvl2 = lvl1.child(j)
                for k in range(lvl2.childCount()):
                    lvl3 = lvl2.child(k)
                    if lvl3.checkState(0) == Qt.Checked:
                        data = lvl3.data(0, Qt.UserRole)
                        checked_items.append((lvl3.text(0), data))
        return checked_items

    # ── Export ──────────────────────────────────────────────────────────────
    def _export_excel(self):

        selected_tables = self.get_checked_tables()
        if not selected_tables:
            CustomMessageBox(
                title="No Table Selected",
                text="Please select at least one table to export.",
                dialogType=MessageBoxType.Warning
            ).exec()
            return
    
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            "results.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        GREEN_HEX  = "FF90AF13"
        WHITE_HEX  = "FFFFFFFF"
        ALT_HEX    = "FFF4F9E3"
        BORDER_HEX = "FF90AF13"

        header_fill  = PatternFill("solid", fgColor=GREEN_HEX)
        alt_fill     = PatternFill("solid", fgColor=ALT_HEX)
        white_fill   = PatternFill("solid", fgColor=WHITE_HEX)
        header_font  = Font(bold=True,  color="FFFFFFFF", name="Calibri", size=11)
        data_font    = Font(bold=False, color="FF2d2d2d", name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_side   = Side(style="thin", color=BORDER_HEX)
        cell_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side,  bottom=thin_side
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for table_name, data in selected_tables:
            columns = data.get("columns", [])
            rows    = data.get("rows",    [])

            if not columns or not rows:
                continue

            ws = wb.create_sheet(title=table_name[:31])

            col_widths = [len(str(col)) for col in columns]
            for row in rows:
                for c_idx, val in enumerate(row):
                    if c_idx < len(col_widths):
                        col_widths[c_idx] = max(col_widths[c_idx], len(str(val)))

            for c_idx, width in enumerate(col_widths):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(c_idx + 1)
                ].width = min(max(width + 4, 14), 50)


            for c_idx, label in enumerate(columns, start=1):
                cell           = ws.cell(row=1, column=c_idx, value=str(label))
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = cell_border
            ws.row_dimensions[1].height = 28

            for r_idx, row in enumerate(rows):
                excel_row  = r_idx + 2
                fill_style = alt_fill if r_idx % 2 == 0 else white_fill
                for c_idx, val in enumerate(row, start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=val)
                    cell.fill      = fill_style
                    cell.font      = data_font
                    cell.alignment = center_align
                    cell.border    = cell_border
                ws.row_dimensions[excel_row].height = 18

        wb.save(path)


class GenerateResultsDialog(QDialog):
    """
    Single frameless dialog with two views:
      Page 0 — Generate Results (table selection)
      Page 1 — Export Table    (preview + export)
    Navigation is handled by a QStackedWidget.
    """

    def __init__(self, parent=None, output_dict: dict = None):
        super().__init__(parent)
        self._output_dict = output_dict or {}
        self.setMinimumWidth(1080)
        self.setMinimumHeight(720)
        self.setObjectName("generate_results_dialog")

        self.setStyleSheet("""
            QDialog#generate_results_dialog {
                background: white;
                border: 1px solid #90AF13;
            }
        """)

        self._setup_ui()

    # ── Wrapper ─────────────────────────────────────────────────────────────
    def _setup_wrapper(self):
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowSystemMenuHint)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(1, 1, 1, 1)
        main_layout.setSpacing(0)

        self.title_bar = CustomTitleBar()
        self.title_bar.setTitle("Generate Results Table")
        main_layout.addWidget(self.title_bar)

        self.content_widget = QWidget(self)
        main_layout.addWidget(self.content_widget, 1)

        size_grip = QSizeGrip(self)
        size_grip.setFixedSize(16, 16)

        overlay = QHBoxLayout()
        overlay.setContentsMargins(0, 0, 4, 4)
        overlay.addStretch()
        overlay.addWidget(size_grip, 0, Qt.AlignBottom | Qt.AlignRight)

        main_layout.addLayout(overlay)

    # ── Main UI ─────────────────────────────────────────────────────────────
    def _setup_ui(self):
        self._setup_wrapper()

        content_layout = QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)

        self.stack = QStackedWidget()
        content_layout.addWidget(self.stack)

        self.page_generate = GenerateResultsPage(
            on_show_selections=self._handle_show_selections
        )
        self.page_generate._cancel_btn.clicked.connect(self.reject)

        self.page_export = ExportTablePage(
            on_back=self._handle_back,
            on_cancel=self.reject
        )

        self.stack.addWidget(self.page_generate)   # index 0
        self.stack.addWidget(self.page_export)     # index 1

        self.stack.setCurrentIndex(0)

    def _get_live_defaults(self) -> dict:
        """
        Deep-copies GENERATE_RESULTS_DEFAULTS and overlays resolved
        (live) table data for any tables that have a resolver.
        Tables without a resolver keep their placeholder data.
        """
        data = copy.deepcopy(GENERATE_RESULTS_DEFAULTS)

        for lvl1_val in data.values():
            for sub_key, sub_val in lvl1_val.items():
                if sub_key in ("id", "label"):
                    continue
                for table_key, table_data in sub_val.items():
                    if table_key in ("id", "label"):
                        continue
                    resolved = resolve_table(
                        table_data["id"],
                        self._output_dict,
                    )
                    if resolved is not None:
                        table_data.update(resolved)

        return data

    # ── Navigation ──────────────────────────────────────────────────────────
    def _handle_show_selections(self):
        selected_names = self.page_generate.get_selected_tables()

        if not selected_names:
            CustomMessageBox(
                title="No Table Selected",
                text="Please select at least one table to continue.",
                dialogType=MessageBoxType.Warning
            ).exec()
            return

        live_defaults = self._get_live_defaults()
        export_data = {}

        for main_key, main_val in live_defaults .items():
            main_bucket = {}

            for sub_key, sub_val in main_val.items():
                if sub_key in ("id", "label"):
                    continue

                group_bucket = {}

                for table_key, table_data in sub_val.items():
                    if table_key in ("id", "label"):
                        continue
                    if table_data["label"] in selected_names:
                        group_bucket[table_data["label"]] = table_data

                if group_bucket:
                    main_bucket[sub_val["label"]] = group_bucket

            if main_bucket:
                export_data[main_val["label"]] = main_bucket

        self.title_bar.setTitle("Export Results")
        self.page_export.load_data(export_data)
        self.stack.setCurrentIndex(1)

    def _handle_back(self):
        self.title_bar.setTitle("Generate Results Table")
        self.stack.setCurrentIndex(0)